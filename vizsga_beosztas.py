#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vizsga Beosztás Szoftver
========================
Egy vizsgabizottság, egy vizsgaterem (max 5 fő), optimális ütemezés.
Többnapos beosztás automatikus nap-felosztással.

Szabályok:
- 1 vizsga = 15 perc, vizsgák között 2 perc szünet a bizottságnak
- Nyelvi tárgyak: 0 perc felkészülési idő
- Egyéb tárgyak: minimum 20 perc felkészülési idő
- Tanuló bent marad a teremben, amíg minden vizsgáját le nem teszi
- Max 5 tanuló lehet a teremben egyszerre
- Cél: bizottság folyamatosan vizsgáztasson ÉS igazságos várakozási idők

Optimalizálási stratégia:
1. Nyelvi tárgyak közbeékelése a nem-nyelviek közé (lyuktöltő szerep)
2. Nem-nyelvi készek közül a legtovább váró vizsgázik (kiegyenlítés)
3. Belépési sorrend:
   - ≤10 vizsgázó: brute-force (az összes permutáció kipróbálása)
   - >10 vizsgázó: gördülő brute-force (lokálisan optimális döntések)
4. Többnapos beosztás: napi időkorlát alapján automatikus felosztás
"""

from dataclasses import dataclass, field
from typing import Optional
from itertools import permutations
import math
import time
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_ELERHETO = True
except ImportError:
    OPENPYXL_ELERHETO = False


# ============================================================
# Adatstruktúrák
# ============================================================

@dataclass
class Targy:
    nev: str
    nyelvi: bool = False

    @property
    def felkeszulesi_ido(self) -> int:
        return 0 if self.nyelvi else 20


@dataclass
class Vizsgazo:
    nev: str
    targyak: list
    hatralevo_targyak: list = field(default_factory=list)
    belepesi_ido: Optional[int] = None
    kilepesi_ido: Optional[int] = None
    felkeszules_vege: Optional[int] = None
    utolso_vizsga_vege: Optional[int] = None
    bent_van: bool = False
    minden_kesz: bool = False
    ossz_varakozas: int = 0

    def __post_init__(self):
        if not self.hatralevo_targyak:
            self.hatralevo_targyak = list(self.targyak)

    @property
    def keszult(self) -> bool:
        return (not self.hatralevo_targyak
                and self.utolso_vizsga_vege is not None)

    def kileptetheto(self, t: int) -> bool:
        return self.keszult and t >= self.utolso_vizsga_vege


@dataclass
class VizsgaEsemeny:
    vizsgazo_nev: str
    targy_nev: str
    targy_nyelvi: bool
    kezdes: int
    befejezes: int
    felkeszules_kezdet: int


# ============================================================
# Tárgy sorrend optimalizálás
# ============================================================

def optimalis_targy_sorrend(targyak: list) -> list:
    """
    Tárgy sorrend:
    - Egy nyelvi előre (belépéskor azonnal vizsgázhat)
    - Nem nyelviek utána
    - Maradék nyelvi tárgyak KÖZBEÉKELVE a nem-nyelviek közé
      → lyuktöltő szerep
    """
    nyelvi = [t for t in targyak if t.nyelvi]
    nem_nyelvi = [t for t in targyak if not t.nyelvi]

    sorrend = []
    if nyelvi:
        sorrend.append(nyelvi.pop(0))

    if not nyelvi:
        sorrend.extend(nem_nyelvi)
    else:
        szamlalo = 0
        for t in nem_nyelvi:
            sorrend.append(t)
            szamlalo += 1
            if szamlalo >= 2 and nyelvi:
                sorrend.append(nyelvi.pop(0))
                szamlalo = 0
        sorrend.extend(nyelvi)

    return sorrend


# ============================================================
# Szimuláció (közös mag)
# ============================================================

VIZSGA_HOSSZ = 15
SZUNET_HOSSZ = 2
MAX_TEREMBEN = 5


def _szimulacio_core(rendezett_vizsgalok: list):
    """
    Eseményvezérelt szimuláció adott belépési sorrenddel.
    Visszatér: (esemenyek, vizsgalok_lista)
    """

    for v in rendezett_vizsgalok:
        v.hatralevo_targyak = list(v.targyak)
        v.bent_van = False
        v.minden_kesz = False
        v.belepesi_ido = None
        v.kilepesi_ido = None
        v.felkeszules_vege = None
        v.utolso_vizsga_vege = None
        v.ossz_varakozas = 0

    esemenyek = []
    varakozok = list(rendezett_vizsgalok)
    bent_levok = []
    bizottsag_szabad = 0
    ido = 0

    def beleptet(t):
        while len(bent_levok) < MAX_TEREMBEN and varakozok:
            v = varakozok.pop(0)
            v.bent_van = True
            v.belepesi_ido = t
            elso = v.hatralevo_targyak[0]
            v.felkeszules_vege = t + elso.felkeszulesi_ido
            bent_levok.append(v)

    def kileptet(t):
        kilep = [v for v in bent_levok if v.kileptetheto(t)]
        for v in kilep:
            v.bent_van = False
            v.minden_kesz = True
            v.kilepesi_ido = t
            bent_levok.remove(v)

    def valassz_vizsgaiot(keszek, t):
        nem_nyelvi = [v for v in keszek if not v.hatralevo_targyak[0].nyelvi]
        nyelvi = [v for v in keszek if v.hatralevo_targyak[0].nyelvi]

        if nem_nyelvi:
            nem_nyelvi.sort(key=lambda v: (
                -(t - v.felkeszules_vege),
                -v.ossz_varakozas,
            ))
            return nem_nyelvi[0]

        if nyelvi:
            nyelvi.sort(key=lambda v: -v.ossz_varakozas)
            return nyelvi[0]

        return None

    beleptet(0)

    for _ in range(10000):
        kileptet(ido)
        beleptet(ido)

        if ido >= bizottsag_szabad:
            keszek = [
                v for v in bent_levok
                if v.hatralevo_targyak
                and v.felkeszules_vege is not None
                and v.felkeszules_vege <= ido
            ]

            if keszek:
                vizsgazo = valassz_vizsgaiot(keszek, ido)
                if vizsgazo is None:
                    vizsgazo = keszek[0]

                targy = vizsgazo.hatralevo_targyak.pop(0)
                vizsga_kezd = ido
                vizsga_veg = ido + VIZSGA_HOSSZ
                felk_kezdet = vizsgazo.felkeszules_vege - targy.felkeszulesi_ido

                if not targy.nyelvi:
                    var = ido - vizsgazo.felkeszules_vege
                    if var > 0:
                        vizsgazo.ossz_varakozas += var

                esemenyek.append(VizsgaEsemeny(
                    vizsgazo_nev=vizsgazo.nev,
                    targy_nev=targy.nev,
                    targy_nyelvi=targy.nyelvi,
                    kezdes=vizsga_kezd,
                    befejezes=vizsga_veg,
                    felkeszules_kezdet=felk_kezdet
                ))

                vizsgazo.utolso_vizsga_vege = vizsga_veg
                bizottsag_szabad = vizsga_veg + SZUNET_HOSSZ

                if vizsgazo.hatralevo_targyak:
                    kov = vizsgazo.hatralevo_targyak[0]
                    vizsgazo.felkeszules_vege = vizsga_veg + kov.felkeszulesi_ido
                else:
                    vizsgazo.felkeszules_vege = None

                continue

        if not bent_levok and not varakozok:
            break

        kovetkezo = []
        if bizottsag_szabad > ido:
            kovetkezo.append(bizottsag_szabad)
        for v in bent_levok:
            if v.felkeszules_vege is not None and v.felkeszules_vege > ido:
                kovetkezo.append(v.felkeszules_vege)
            if v.utolso_vizsga_vege is not None and v.utolso_vizsga_vege > ido:
                kovetkezo.append(v.utolso_vizsga_vege)

        if kovetkezo:
            uj = min(kovetkezo)
            ido = max(uj, ido + 1)
        else:
            ido += 1

    return esemenyek, rendezett_vizsgalok


def _ertekeles(esemenyek, vizsgalok):
    """Pontozás: (összes_idő, max_várakozás, üresjárat). Kisebb = jobb."""
    if not esemenyek:
        return (99999, 99999, 99999)

    ossz_targy = sum(len(v.targyak) for v in vizsgalok)
    if len(esemenyek) != ossz_targy:
        return (99999, 99999, 99999)

    elso = min(e.kezdes for e in esemenyek)
    utolso = max(e.befejezes for e in esemenyek)
    teljes = utolso - elso

    elfoglalt = 0
    for i, e in enumerate(esemenyek):
        elfoglalt += VIZSGA_HOSSZ
        if i < len(esemenyek) - 1:
            szunet = esemenyek[i + 1].kezdes - e.befejezes
            elfoglalt += min(szunet, SZUNET_HOSSZ)
    uresjarat = teljes - elfoglalt

    max_var = 0
    for v in vizsgalok:
        for e in esemenyek:
            if e.vizsgazo_nev == v.nev and not e.targy_nyelvi:
                var = e.kezdes - (e.felkeszules_kezdet + 20)
                if var > max_var:
                    max_var = var

    return (teljes, max_var, uresjarat)


# ============================================================
# Stratégia 1: Teljes brute-force (≤10 vizsgázó)
# ============================================================

def _brute_force_szimulacio(vizsgalok):
    """Minden permutációt kipróbál, a legjobbat adja vissza."""
    n = len(vizsgalok)
    fakt = math.factorial(n)
    print(f"  Stratégia: TELJES BRUTE-FORCE ({n} vizsgázó, {fakt:,} permutáció)")

    legjobb_ertek = (99999, 99999, 99999)
    legjobb_sorrend = None

    for perm in permutations(range(n)):
        rendezett = [vizsgalok[i] for i in perm]
        esemenyek, _ = _szimulacio_core(rendezett)
        ertek = _ertekeles(esemenyek, vizsgalok)
        if ertek < legjobb_ertek:
            legjobb_ertek = ertek
            legjobb_sorrend = list(perm)

    rendezett = [vizsgalok[i] for i in legjobb_sorrend]
    return _szimulacio_core(rendezett)


# ============================================================
# Stratégia 2: Gördülő brute-force (>10 vizsgázó)
# ============================================================

def _gordulo_szimulacio(vizsgalok):
    """
    Gördülő brute-force: az első 5 belépőre teljes keresés,
    utána pozíciónként optimalizálás.
    """
    n = len(vizsgalok)
    print(f"  Stratégia: GÖRDÜLŐ BRUTE-FORCE ({n} vizsgázó)")

    jeloltek = sorted(vizsgalok, key=lambda v: (
        0 if v.hatralevo_targyak and v.hatralevo_targyak[0].nyelvi else 1,
        -len(v.targyak),
    ))

    max_jelolt = min(len(jeloltek), 8)
    elso_jeloltek = jeloltek[:max_jelolt]
    tobbi = [v for v in vizsgalok if v not in elso_jeloltek]

    elso_n = min(MAX_TEREMBEN, len(elso_jeloltek))
    p_count = math.perm(max_jelolt, elso_n)
    print(f"  Első {elso_n} hely: {max_jelolt} jelöltből → {p_count:,} variáció")

    legjobb_ertek = (99999, 99999, 99999)
    legjobb_sorrend = None
    legjobb_teljes = None

    elso_indexek = list(range(max_jelolt))

    for elso_perm in permutations(elso_indexek, elso_n):
        elso_ot = [elso_jeloltek[i] for i in elso_perm]
        maradek_jeloltek = [elso_jeloltek[i] for i in range(max_jelolt) if i not in elso_perm]
        maradek = maradek_jeloltek + tobbi
        maradek.sort(key=lambda v: (
            0 if v.hatralevo_targyak and v.hatralevo_targyak[0].nyelvi else 1,
            -len(v.targyak),
        ))
        teljes_sorrend = elso_ot + maradek

        esemenyek, _ = _szimulacio_core(teljes_sorrend)
        ertek = _ertekeles(esemenyek, vizsgalok)
        if ertek < legjobb_ertek:
            legjobb_ertek = ertek
            legjobb_sorrend = [v.nev for v in teljes_sorrend]
            legjobb_teljes = list(teljes_sorrend)

    print(f"  Első kör legjobb: {legjobb_ertek[0]}p, {legjobb_ertek[1]}p max vár, {legjobb_ertek[2]}p üresj.")
    print(f"  Maradék belépési sorrend finomítása...")

    elso_ot_nevek = legjobb_sorrend[:elso_n]
    maradek_vizsgalok = [v for v in legjobb_teljes if v.nev not in elso_ot_nevek]
    elso_ot_obj = [v for v in legjobb_teljes if v.nev in elso_ot_nevek]
    elso_ot_obj.sort(key=lambda v: elso_ot_nevek.index(v.nev))

    vegleges_maradek = []
    meg_nem_beleptetve = list(maradek_vizsgalok)

    for pozicio in range(len(maradek_vizsgalok)):
        legjobb_jelolt = None
        legjobb_jelolt_ertek = (99999, 99999, 99999)

        for j, jelolt in enumerate(meg_nem_beleptetve):
            proba_maradek = [jelolt] + [v for v in meg_nem_beleptetve if v is not jelolt]
            proba_sorrend = elso_ot_obj + vegleges_maradek + proba_maradek

            esemenyek, _ = _szimulacio_core(proba_sorrend)
            ertek = _ertekeles(esemenyek, vizsgalok)

            if ertek < legjobb_jelolt_ertek:
                legjobb_jelolt_ertek = ertek
                legjobb_jelolt = jelolt

        vegleges_maradek.append(legjobb_jelolt)
        meg_nem_beleptetve.remove(legjobb_jelolt)

    vegleges_sorrend = elso_ot_obj + vegleges_maradek
    return _szimulacio_core(vegleges_sorrend)


# ============================================================
# Fő szimuláció (automatikus stratégia-választás)
# ============================================================

BRUTE_FORCE_LIMIT = 10


def szimulacio(vizsgalok: list):
    """Optimális ütemezés: automatikusan választ stratégiát a létszám alapján."""

    for v in vizsgalok:
        v.targyak = optimalis_targy_sorrend(v.targyak)

    n = len(vizsgalok)
    start = time.time()

    print(f"  Vizsgázók száma: {n}")
    print(f"  Összes vizsga:   {sum(len(v.targyak) for v in vizsgalok)}")
    print()

    if n <= BRUTE_FORCE_LIMIT:
        esemenyek, vizsgalok_eredmeny = _brute_force_szimulacio(vizsgalok)
    else:
        esemenyek, vizsgalok_eredmeny = _gordulo_szimulacio(vizsgalok)

    elapsed = time.time() - start
    ertek = _ertekeles(esemenyek, vizsgalok)

    print()
    print(f"  Eredmény: {ertek[0]}p összes, {ertek[1]}p max várakozás, {ertek[2]}p üresjárat")
    print(f"  Belépési sorrend: {' → '.join(v.nev for v in vizsgalok_eredmeny)}")
    print(f"  Futásidő: {elapsed:.1f} mp")
    print()

    return esemenyek, vizsgalok_eredmeny


# ============================================================
# Nap-felosztás
# ============================================================

def napokra_oszt(vizsgalok: list, napi_perc: int) -> list:
    """
    Vizsgázókat napokra osztja a rendelkezésre álló idő alapján.

    Egy nap max ennyi vizsgát tud befogadni:
      max_vizsga = (napi_perc + SZUNET_HOSSZ) // (VIZSGA_HOSSZ + SZUNET_HOSSZ)

    Stratégia:
    1. Kiszámolja hány nap kell minimum
    2. A vizsgákat egyenletesen elosztja a napok között
    3. A vizsgázókat úgy rendezi, hogy minden napon legyen
       nyelvi tárggyal kezdő (hatékony indítás)
    """
    max_vizsga_per_nap = (napi_perc + SZUNET_HOSSZ) // (VIZSGA_HOSSZ + SZUNET_HOSSZ)

    ossz_vizsga = sum(len(v.targyak) for v in vizsgalok)
    min_napok = max(1, math.ceil(ossz_vizsga / max_vizsga_per_nap))

    # Rendezés: legtöbb tárgy előre, nyelvi tárggyal kezdők előnyben
    rendezett = sorted(vizsgalok, key=lambda v: (
        -len(v.targyak),
        0 if any(t.nyelvi for t in v.targyak) else 1,
    ))

    # Elosztás: round-robin a napok között, vizsga-szám alapján egyenlítve
    napok = [[] for _ in range(min_napok)]
    nap_vizsgak = [0] * min_napok

    for v in rendezett:
        targy_szam = len(v.targyak)
        # A legkevesebb vizsgával rendelkező napra tesszük, ha belefér
        legjobb_nap = None
        legjobb_szam = 99999
        for i in range(min_napok):
            if nap_vizsgak[i] + targy_szam <= max_vizsga_per_nap:
                if nap_vizsgak[i] < legjobb_szam:
                    legjobb_szam = nap_vizsgak[i]
                    legjobb_nap = i

        # Ha nem fér sehova, de új nap nyitása pazarló lenne,
        # engedjük be a legkevesebb vizsgával rendelkező napra
        if legjobb_nap is None and len(napok) == min_napok:
            legkevesebb_idx = min(range(min_napok), key=lambda i: nap_vizsgak[i])
            legjobb_nap = legkevesebb_idx

        if legjobb_nap is not None:
            napok[legjobb_nap].append(v)
            nap_vizsgak[legjobb_nap] += targy_szam
        else:
            # Nem fér be sehova → új nap kell
            napok.append([v])
            nap_vizsgak.append(targy_szam)

    # Üres napok eltávolítása
    napok = [n for n in napok if n]

    return napok, max_vizsga_per_nap


# ============================================================
# Kimenet
# ============================================================

def perc_to_ido(perc: int, kezdes_ora: int = 8) -> str:
    ora = kezdes_ora + perc // 60
    p = perc % 60
    return f"{ora}:{p:02d}"


def ellenorzes(esemenyek, vizsgalok):
    hibak = []
    if esemenyek:
        utolso = max(e.befejezes for e in esemenyek)
        for p in range(0, utolso + 1):
            cnt = sum(1 for v in vizsgalok
                      if v.belepesi_ido is not None and v.kilepesi_ido is not None
                      and v.belepesi_ido <= p < v.kilepesi_ido)
            if cnt > MAX_TEREMBEN:
                hibak.append(f"  ⚠ t={perc_to_ido(p)}: {cnt} fő (max {MAX_TEREMBEN})")

    for i in range(len(esemenyek) - 1):
        e1, e2 = esemenyek[i], esemenyek[i + 1]
        if e2.kezdes - e1.befejezes < SZUNET_HOSSZ:
            hibak.append(f"  ⚠ Szünet hiba: {e1.targy_nev} és {e2.targy_nev} között")

    for e in esemenyek:
        felk = e.kezdes - e.felkeszules_kezdet
        if not e.targy_nyelvi and felk < 20:
            hibak.append(f"  ⚠ Felkészülés hiba: {e.vizsgazo_nev}/{e.targy_nev} ({felk}p < 20p)")

    ossz_targy = sum(len(v.targyak) for v in vizsgalok)
    if len(esemenyek) != ossz_targy:
        hibak.append(f"  ⚠ Hiányzó vizsgák: {ossz_targy} tárgy, de csak {len(esemenyek)} vizsga")

    return hibak


def kiiras_nap(esemenyek, vizsgalok, nap_szam: int, kezdes_ora: int = 8):
    """Egyetlen nap beosztásának kiírása."""

    print(f"  ┌─────────────────────────────────────────────────────────────────────────────┐")
    print(f"  │  {nap_szam}. NAP{' ' * 71}│")
    print(f"  └─────────────────────────────────────────────────────────────────────────────┘")
    print()

    if esemenyek:
        elso_kezd = min(e.kezdes for e in esemenyek)
        utolso_veg = max(e.befejezes for e in esemenyek)
        teljes = utolso_veg - elso_kezd
        ossz_targy = sum(len(v.targyak) for v in vizsgalok)
        print(f"  Vizsga kezdete:     {perc_to_ido(elso_kezd, kezdes_ora)}")
        print(f"  Utolsó vizsga vége: {perc_to_ido(utolso_veg, kezdes_ora)}")
        print(f"  Összes időtartam:   {teljes} perc ({teljes // 60} óra {teljes % 60} perc)")
        print(f"  Vizsgák száma:      {len(esemenyek)}")
        print(f"  Vizsgázók száma:    {len(vizsgalok)} (összesen {ossz_targy} tárgy)")
        print()

    # Bizottság idővonala
    print("-" * 80)
    print("  BIZOTTSÁG IDŐVONALA")
    print("-" * 80)
    for i, e in enumerate(esemenyek, 1):
        ny = " [NY]" if e.targy_nyelvi else ""
        print(f"  {i:2d}. {perc_to_ido(e.kezdes, kezdes_ora)}-{perc_to_ido(e.befejezes, kezdes_ora)}"
              f"  {e.vizsgazo_nev:<18s}  {e.targy_nev}{ny}")
        if i < len(esemenyek):
            kov = esemenyek[i]
            szunet = kov.kezdes - e.befejezes
            if szunet > SZUNET_HOSSZ:
                print(f"       ⚠  {szunet}p várakozás ({SZUNET_HOSSZ}p szünet"
                      f" + {szunet - SZUNET_HOSSZ}p üresjárat)")
    print()

    # Vizsgázók
    print("-" * 80)
    print("  VIZSGÁZÓK RÉSZLETES BEOSZTÁSA")
    print("-" * 80)
    for v in vizsgalok:
        v_esem = [e for e in esemenyek if e.vizsgazo_nev == v.nev]
        if not v_esem:
            print(f"\n  {v.nev}: NEM ÜTEMEZETT!")
            continue

        var_lista = []
        for e in v_esem:
            if not e.targy_nyelvi:
                var = e.kezdes - (e.felkeszules_kezdet + 20)
                if var > 0:
                    var_lista.append(var)

        max_var = max(var_lista) if var_lista else 0
        atl_var = sum(var_lista) / len(var_lista) if var_lista else 0

        print(f"\n  {v.nev}")
        print(f"  {'─' * 75}")
        print(f"    Belépés a terembe: {perc_to_ido(v.belepesi_ido, kezdes_ora)}")
        print(f"    Kilépés:           {perc_to_ido(v.kilepesi_ido, kezdes_ora)}")
        print(f"    Bent töltött idő:  {v.kilepesi_ido - v.belepesi_ido} perc")
        print(f"    Tárgyak száma:     {len(v.targyak)}")
        if var_lista:
            print(f"    Várakozás:         max {max_var}p, átlag {atl_var:.0f}p")
        print()
        print(f"    {'Tárgy':<20s}  {'Felkészülés':<20s}  {'Vizsga':<15s}  Megjegyzés")
        print(f"    {'─' * 20}  {'─' * 20}  {'─' * 15}  {'─' * 20}")
        for e in v_esem:
            vizsga_str = f"{perc_to_ido(e.kezdes, kezdes_ora)}-{perc_to_ido(e.befejezes, kezdes_ora)}"
            if e.targy_nyelvi:
                felk_str = "– (nyelvi)"
                megj = "nincs felkészülés [NY]"
            else:
                felk_str = (f"{perc_to_ido(e.felkeszules_kezdet, kezdes_ora)}-"
                            f"{perc_to_ido(e.felkeszules_kezdet + 20, kezdes_ora)}")
                varakozas = e.kezdes - (e.felkeszules_kezdet + 20)
                if varakozas > 0:
                    megj = f"felk. 20p, vár {varakozas}p"
                else:
                    megj = "felk. 20p"
            print(f"    {e.targy_nev:<20s}  {felk_str:<20s}  {vizsga_str:<15s}  {megj}")
    print()

    # Terem
    print("-" * 80)
    print("  TEREM KIHASZNÁLTSÁG")
    print("-" * 80)
    if esemenyek:
        utolso = max(e.befejezes for e in esemenyek)
        letszam = {}
        for p in range(0, utolso + 1):
            letszam[p] = sum(1 for v in vizsgalok
                             if v.belepesi_ido is not None and v.kilepesi_ido is not None
                             and v.belepesi_ido <= p < v.kilepesi_ido)

        max_l = max(letszam.values()) if letszam else 0
        print(f"  Max létszám: {max_l} fő (limit: {MAX_TEREMBEN})")
        print()
        print("  Idő      Fő  Terem")
        print("  " + "─" * 35)
        for p in range(0, utolso + 15, 15):
            l = letszam.get(p, 0)
            bar = "█" * l + "░" * (MAX_TEREMBEN - l)
            print(f"  {perc_to_ido(p, kezdes_ora):>6s}   {l}   [{bar}]")
    print()

    # Statisztika
    print("-" * 80)
    print("  STATISZTIKA")
    print("-" * 80)
    if esemenyek:
        ossz_vizsga = len(esemenyek) * VIZSGA_HOSSZ
        elso = min(e.kezdes for e in esemenyek)
        utolso = max(e.befejezes for e in esemenyek)
        teljes = utolso - elso
        elfoglalt = 0
        for i, e in enumerate(esemenyek):
            elfoglalt += VIZSGA_HOSSZ
            if i < len(esemenyek) - 1:
                szunet = esemenyek[i + 1].kezdes - e.befejezes
                elfoglalt += min(szunet, SZUNET_HOSSZ)
        uresjarat = teljes - elfoglalt
        print(f"  Vizsgáztatás:   {ossz_vizsga}p ({ossz_vizsga // 60}h {ossz_vizsga % 60}p)")
        print(f"  Köt. szünetek:  {(len(esemenyek) - 1) * SZUNET_HOSSZ}p")
        print(f"  Üresjárat:      {uresjarat}p")
        if teljes > 0:
            print(f"  Kihasználtság:  {ossz_vizsga / teljes * 100:.1f}%")

        ossz_var = []
        for v in vizsgalok:
            for e in [ev for ev in esemenyek if ev.vizsgazo_nev == v.nev]:
                if not e.targy_nyelvi:
                    var = e.kezdes - (e.felkeszules_kezdet + 20)
                    if var > 0:
                        ossz_var.append(var)
        if ossz_var:
            print()
            print(f"  Várakozás (felk. feletti):")
            print(f"    Max:    {max(ossz_var)}p")
            print(f"    Átlag:  {sum(ossz_var) / len(ossz_var):.1f}p")
            print(f"    Összeg: {sum(ossz_var)}p")

    print()
    hibak = ellenorzes(esemenyek, vizsgalok)
    if hibak:
        print("-" * 80)
        print("  HIBÁK")
        print("-" * 80)
        for h in hibak[:15]:
            print(h)
        if len(hibak) > 15:
            print(f"  ... és {len(hibak) - 15} további hiba")
    else:
        print("  ✓ Minden szabály teljesül.")

    print()


# ============================================================
# Excel export
# ============================================================

def excel_export(nap_adatok: list, kezdes_ora: int, veg_ora: int, fajlnev: str = "vizsga_beosztas.xlsx"):
    """
    Excel fájl generálása 3 munkalappal:
      1) Statisztika — összesítő adatok
      2) Bizottság beosztása — naponkénti idővonal
      3) Diákok beosztása — naponkénti részletes beosztás

    nap_adatok: [(esemenyek, vizsgalok, ertek), ...] naponként
    """
    if not OPENPYXL_ELERHETO:
        print("  ⚠ openpyxl nem elérhető, Excel export kihagyva.")
        print("  Telepítés: pip install openpyxl")
        return None

    wb = Workbook()

    # --- Stílusok ---
    cim_font = Font(name="Calibri", size=14, bold=True)
    alcim_font = Font(name="Calibri", size=11, bold=True)
    fejlec_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fejlec_fill = PatternFill(start_color="01696F", end_color="01696F", fill_type="solid")
    nap_fill = PatternFill(start_color="E8F5F6", end_color="E8F5F6", fill_type="solid")
    nap_font = Font(name="Calibri", size=11, bold=True, color="01696F")
    normal_font = Font(name="Calibri", size=10)
    nyelvi_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    szam_font = Font(name="Calibri", size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="D4D1CA")
    )
    ertek_font = Font(name="Calibri", size=12, bold=True)
    jo_font = Font(name="Calibri", size=10, color="437A22")

    def fejlec_sor(ws, sor, oszlopok, start_col=1):
        for j, nev in enumerate(oszlopok, start_col):
            c = ws.cell(row=sor, column=j, value=nev)
            c.font = fejlec_font
            c.fill = fejlec_fill
            c.alignment = Alignment(horizontal="center", vertical="center")

    def auto_szelesseg(ws, min_w=10, max_w=30):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

    ossz_vizsgak = sum(len(e) for e, _, _ in nap_adatok)
    ossz_targyak = sum(sum(len(v.targyak) for v in vizs) for _, vizs, _ in nap_adatok)
    ossz_vizsgalok = sum(len(vizs) for _, vizs, _ in nap_adatok)
    ossz_ido = sum(ert[0] for _, _, ert in nap_adatok)
    ossz_uresjarat = sum(ert[2] for _, _, ert in nap_adatok)
    ossz_max_var = max(ert[1] for _, _, ert in nap_adatok)
    napi_perc = (veg_ora - kezdes_ora) * 60

    # ================================================================
    # 1. MUNKALAP: Statisztika
    # ================================================================
    ws_stat = wb.active
    ws_stat.title = "Statisztika"
    ws_stat.sheet_properties.tabColor = "01696F"

    ws_stat.column_dimensions['A'].width = 3

    r = 2
    ws_stat.cell(row=r, column=2, value="Vizsga Beosztás — Statisztika").font = cim_font
    ws_stat.merge_cells('B2:E2')
    r += 1
    ws_stat.cell(row=r, column=2,
                 value=f"Generálva: {datetime.now().strftime('%Y.%m.%d %H:%M')}").font = Font(
                     name="Calibri", size=9, color="7A7974")
    r += 2

    # Összesítő blokk
    ws_stat.cell(row=r, column=2, value="Összesítés").font = alcim_font
    r += 1
    stat_adatok = [
        ("Napok száma", len(nap_adatok)),
        ("Vizsgázók száma", ossz_vizsgalok),
        ("Összes vizsga", ossz_targyak),
        ("Napi időkeret", f"{kezdes_ora}:00 – {veg_ora}:00 ({napi_perc} perc)"),
        ("", ""),
        ("Összes vizsgaidő", f"{ossz_ido} perc ({ossz_ido // 60}h {ossz_ido % 60}p)"),
        ("Összes üresjárat", f"{ossz_uresjarat} perc"),
        ("Max várakozás", f"{ossz_max_var} perc"),
        ("Kihasználtság", f"{ossz_targyak * VIZSGA_HOSSZ / ossz_ido * 100:.1f}%"),
    ]
    for nev, ertek in stat_adatok:
        if nev == "":
            r += 1
            continue
        ws_stat.cell(row=r, column=2, value=nev).font = normal_font
        c = ws_stat.cell(row=r, column=4, value=ertek)
        c.font = ertek_font if nev in ("Kihasználtság", "Összes üresjárat", "Max várakozás") else normal_font
        if nev == "Összes üresjárat" and ossz_uresjarat <= 5:
            c.font = Font(name="Calibri", size=12, bold=True, color="437A22")
        if nev == "Max várakozás" and ossz_max_var <= 20:
            c.font = Font(name="Calibri", size=12, bold=True, color="437A22")
        r += 1

    r += 2

    # Napi bontás táblázat
    ws_stat.cell(row=r, column=2, value="Napi bontás").font = alcim_font
    r += 1
    fejlec_sor(ws_stat, r, ["Nap", "Vizsgázók", "Vizsgák", "Időtartam",
                             "Üresjárat", "Max vár.", "Kezdés", "Befejezés"], 2)
    r += 1
    for i, (esem, vizs, ert) in enumerate(nap_adatok, 1):
        elso_k = min(e.kezdes for e in esem) if esem else 0
        utolso_v = max(e.befejezes for e in esem) if esem else 0
        ws_stat.cell(row=r, column=2, value=f"{i}. nap").font = normal_font
        ws_stat.cell(row=r, column=3, value=len(vizs)).font = szam_font
        ws_stat.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws_stat.cell(row=r, column=4, value=len(esem)).font = szam_font
        ws_stat.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        ws_stat.cell(row=r, column=5, value=f"{ert[0]} perc").font = szam_font
        ws_stat.cell(row=r, column=6, value=f"{ert[2]} perc").font = szam_font
        if ert[2] <= 5:
            ws_stat.cell(row=r, column=6).font = jo_font
        ws_stat.cell(row=r, column=7, value=f"{ert[1]} perc").font = szam_font
        if ert[1] <= 20:
            ws_stat.cell(row=r, column=7).font = jo_font
        ws_stat.cell(row=r, column=8, value=perc_to_ido(elso_k, kezdes_ora)).font = szam_font
        ws_stat.cell(row=r, column=8).alignment = Alignment(horizontal="center")
        ws_stat.cell(row=r, column=9, value=perc_to_ido(utolso_v, kezdes_ora)).font = szam_font
        ws_stat.cell(row=r, column=9).alignment = Alignment(horizontal="center")
        for col in range(2, 10):
            ws_stat.cell(row=r, column=col).border = thin_border
        r += 1

    r += 2
    # Vizsgázók listája naponként
    ws_stat.cell(row=r, column=2, value="Vizsgázók napokra beosztva").font = alcim_font
    r += 1
    for i, (esem, vizs, ert) in enumerate(nap_adatok, 1):
        ws_stat.cell(row=r, column=2, value=f"{i}. nap:").font = nap_font
        ws_stat.cell(row=r, column=3, value=", ".join(v.nev for v in vizs)).font = normal_font
        ws_stat.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        r += 1

    ws_stat.column_dimensions['B'].width = 22
    ws_stat.column_dimensions['C'].width = 14
    ws_stat.column_dimensions['D'].width = 14
    ws_stat.column_dimensions['E'].width = 16
    ws_stat.column_dimensions['F'].width = 14
    ws_stat.column_dimensions['G'].width = 14
    ws_stat.column_dimensions['H'].width = 12
    ws_stat.column_dimensions['I'].width = 12

    # ================================================================
    # 2. MUNKALAP: Bizottság beosztása
    # ================================================================
    ws_biz = wb.create_sheet("Bizottság beosztása")
    ws_biz.sheet_properties.tabColor = "A84B2F"
    ws_biz.column_dimensions['A'].width = 3

    r = 2
    ws_biz.cell(row=r, column=2, value="Bizottság beosztása").font = cim_font
    ws_biz.merge_cells('B2:H2')
    r += 2

    for nap_idx, (esem, vizs, ert) in enumerate(nap_adatok, 1):
        # Nap fejléc
        ws_biz.cell(row=r, column=2, value=f"{nap_idx}. nap").font = nap_font
        ws_biz.cell(row=r, column=2).fill = nap_fill
        for col in range(2, 9):
            ws_biz.cell(row=r, column=col).fill = nap_fill
        r += 1

        fejlec_sor(ws_biz, r, ["Ssz.", "Kezdés", "Befejezés", "Vizsgázó",
                                "Tárgy", "Típus", "Megjegyzés"], 2)
        r += 1

        for j, e in enumerate(esem, 1):
            ws_biz.cell(row=r, column=2, value=j).font = szam_font
            ws_biz.cell(row=r, column=2).alignment = Alignment(horizontal="center")
            ws_biz.cell(row=r, column=3, value=perc_to_ido(e.kezdes, kezdes_ora)).font = szam_font
            ws_biz.cell(row=r, column=3).alignment = Alignment(horizontal="center")
            ws_biz.cell(row=r, column=4, value=perc_to_ido(e.befejezes, kezdes_ora)).font = szam_font
            ws_biz.cell(row=r, column=4).alignment = Alignment(horizontal="center")
            ws_biz.cell(row=r, column=5, value=e.vizsgazo_nev).font = normal_font
            ws_biz.cell(row=r, column=6, value=e.targy_nev).font = normal_font
            tipus = "Nyelvi" if e.targy_nyelvi else "Nem nyelvi"
            ws_biz.cell(row=r, column=7, value=tipus).font = szam_font
            ws_biz.cell(row=r, column=7).alignment = Alignment(horizontal="center")

            # Megjegyzés: üresjárat jelzés
            megj = ""
            if j > 1:
                elozo = esem[j - 2]
                szunet = e.kezdes - elozo.befejezes
                if szunet > SZUNET_HOSSZ:
                    megj = f"{szunet - SZUNET_HOSSZ}p üresjárat"
            ws_biz.cell(row=r, column=8, value=megj).font = Font(
                name="Calibri", size=9, color="964219") if megj else normal_font

            if e.targy_nyelvi:
                for col in range(2, 9):
                    ws_biz.cell(row=r, column=col).fill = nyelvi_fill

            for col in range(2, 9):
                ws_biz.cell(row=r, column=col).border = thin_border
            r += 1

        r += 2  # Napok között üres sorok

    ws_biz.column_dimensions['B'].width = 7
    ws_biz.column_dimensions['C'].width = 10
    ws_biz.column_dimensions['D'].width = 12
    ws_biz.column_dimensions['E'].width = 20
    ws_biz.column_dimensions['F'].width = 20
    ws_biz.column_dimensions['G'].width = 12
    ws_biz.column_dimensions['H'].width = 18

    # ================================================================
    # 3. MUNKALAP: Diákok beosztása
    # ================================================================
    ws_diak = wb.create_sheet("Diákok beosztása")
    ws_diak.sheet_properties.tabColor = "1B474D"
    ws_diak.column_dimensions['A'].width = 3

    r = 2
    ws_diak.cell(row=r, column=2, value="Diákok beosztása").font = cim_font
    ws_diak.merge_cells('B2:I2')
    r += 2

    for nap_idx, (esem, vizs, ert) in enumerate(nap_adatok, 1):
        # Nap fejléc
        ws_diak.cell(row=r, column=2, value=f"{nap_idx}. nap").font = nap_font
        ws_diak.cell(row=r, column=2).fill = nap_fill
        for col in range(2, 11):
            ws_diak.cell(row=r, column=col).fill = nap_fill
        r += 1

        for v in vizs:
            v_esem = [e for e in esem if e.vizsgazo_nev == v.nev]
            if not v_esem:
                continue

            # Vizsgázó fejléc sor
            ws_diak.cell(row=r, column=2, value=v.nev).font = alcim_font
            bent_ido = (v.kilepesi_ido - v.belepesi_ido
                        if v.kilepesi_ido is not None and v.belepesi_ido is not None
                        else 0)
            ws_diak.cell(row=r, column=5,
                         value=f"Belépés: {perc_to_ido(v.belepesi_ido, kezdes_ora)}"
                         ).font = Font(name="Calibri", size=9, color="7A7974")
            ws_diak.cell(row=r, column=7,
                         value=f"Kilépés: {perc_to_ido(v.kilepesi_ido, kezdes_ora)}"
                         ).font = Font(name="Calibri", size=9, color="7A7974")
            ws_diak.cell(row=r, column=9,
                         value=f"Bent: {bent_ido}p"
                         ).font = Font(name="Calibri", size=9, color="7A7974")
            r += 1

            # Tárgy fejléc
            fejlec_sor(ws_diak, r, ["Tárgy", "Felk. kezdet", "Felk. vége",
                                     "Vizsga kezdet", "Vizsga vége", "Típus",
                                     "Várakozás"], 2)
            r += 1

            for e in v_esem:
                ws_diak.cell(row=r, column=2, value=e.targy_nev).font = normal_font
                if e.targy_nyelvi:
                    ws_diak.cell(row=r, column=3, value="–").font = szam_font
                    ws_diak.cell(row=r, column=3).alignment = Alignment(horizontal="center")
                    ws_diak.cell(row=r, column=4, value="–").font = szam_font
                    ws_diak.cell(row=r, column=4).alignment = Alignment(horizontal="center")
                else:
                    ws_diak.cell(row=r, column=3,
                                 value=perc_to_ido(e.felkeszules_kezdet, kezdes_ora)).font = szam_font
                    ws_diak.cell(row=r, column=3).alignment = Alignment(horizontal="center")
                    ws_diak.cell(row=r, column=4,
                                 value=perc_to_ido(e.felkeszules_kezdet + 20, kezdes_ora)).font = szam_font
                    ws_diak.cell(row=r, column=4).alignment = Alignment(horizontal="center")

                ws_diak.cell(row=r, column=5,
                             value=perc_to_ido(e.kezdes, kezdes_ora)).font = szam_font
                ws_diak.cell(row=r, column=5).alignment = Alignment(horizontal="center")
                ws_diak.cell(row=r, column=6,
                             value=perc_to_ido(e.befejezes, kezdes_ora)).font = szam_font
                ws_diak.cell(row=r, column=6).alignment = Alignment(horizontal="center")

                tipus = "Nyelvi" if e.targy_nyelvi else "Nem nyelvi"
                ws_diak.cell(row=r, column=7, value=tipus).font = szam_font
                ws_diak.cell(row=r, column=7).alignment = Alignment(horizontal="center")

                if e.targy_nyelvi:
                    var_str = "–"
                else:
                    var = e.kezdes - (e.felkeszules_kezdet + 20)
                    var_str = f"{var}p" if var > 0 else "0p"
                ws_diak.cell(row=r, column=8, value=var_str).font = szam_font
                ws_diak.cell(row=r, column=8).alignment = Alignment(horizontal="center")

                if e.targy_nyelvi:
                    for col in range(2, 9):
                        ws_diak.cell(row=r, column=col).fill = nyelvi_fill

                for col in range(2, 9):
                    ws_diak.cell(row=r, column=col).border = thin_border
                r += 1

            r += 1  # Vizsgázók között üres sor

        r += 1  # Napok között extra üres sor

    ws_diak.column_dimensions['B'].width = 20
    ws_diak.column_dimensions['C'].width = 14
    ws_diak.column_dimensions['D'].width = 14
    ws_diak.column_dimensions['E'].width = 14
    ws_diak.column_dimensions['F'].width = 14
    ws_diak.column_dimensions['G'].width = 12
    ws_diak.column_dimensions['H'].width = 12

    # Mentés
    wb.save(fajlnev)
    return fajlnev


# ============================================================
# Excel bemenet
# ============================================================

def beolvas_excelbol(fajlnev: str) -> list:
    """
    Vizsgázók beolvasása Excel fájlból.

    Elvárt struktúra ("Vizsgázók" munkalap):
      A oszlop: Vizsgázó neve
      B-G oszlopok: Tárgyak (max 6, de további oszlopok is elfogadva)
      Nyelvi tárgyak: * jel a név előtt (pl. "*Angol nyelv")
      Üres cellák: kihagyva
      Első sor: fejléc (kihagyva)
    """
    if not OPENPYXL_ELERHETO:
        print("  ⚠ openpyxl nem elérhető, Excel beolvasás nem lehetséges.")
        print("  Telepítés: pip install openpyxl")
        return []

    from openpyxl import load_workbook

    try:
        wb = load_workbook(fajlnev, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"  ⚠ Fájl nem található: {fajlnev}")
        return []
    except Exception as e:
        print(f"  ⚠ Hiba a fájl megnyitásakor: {e}")
        return []

    # Első munkalapot használjuk
    ws = wb.active

    vizsgalok = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue  # Üres sor

        nev = str(row[0]).strip()
        if not nev:
            continue

        targyak = []
        for cella in row[1:]:
            if cella is None:
                continue
            targy_nev = str(cella).strip()
            if not targy_nev:
                continue

            nyelvi = targy_nev.startswith("*")
            if nyelvi:
                targy_nev = targy_nev[1:].strip()

            targyak.append(Targy(nev=targy_nev, nyelvi=nyelvi))

        if targyak:
            vizsgalok.append(Vizsgazo(nev=nev, targyak=targyak))
        # Nincs tárgy → csendben kihagyva (útmutató sorok, megjegyzések stb.)

    wb.close()

    if vizsgalok:
        print(f"  Beolvasva: {len(vizsgalok)} vizsgázó,"
              f" {sum(len(v.targyak) for v in vizsgalok)} tárgy")
        for v in vizsgalok:
            nyelvi_db = sum(1 for t in v.targyak if t.nyelvi)
            print(f"    {v.nev}: {len(v.targyak)} tárgy"
                  f" ({nyelvi_db} nyelvi)" if nyelvi_db else
                  f"    {v.nev}: {len(v.targyak)} tárgy")
    else:
        print("  ⚠ Nem található vizsgázó adat a fájlban.")

    return vizsgalok


# ============================================================
# Felhasználói bemenet
# ============================================================

def beolvas_idokeretet() -> tuple:
    """Bekéri a vizsgaidőszak paramétereit."""
    print()
    print("  ╔══════════════════════════════════════════════════════╗")
    print("  ║            VIZSGA BEOSZTÁS SZOFTVER                 ║")
    print("  ╠══════════════════════════════════════════════════════╣")
    print("  ║  Egy vizsgabizottság, max 5 fő a teremben           ║")
    print("  ║  1 vizsga = 15 perc, 2 perc szünet közöttük         ║")
    print("  ╚══════════════════════════════════════════════════════╝")
    print()

    while True:
        try:
            bev = input("  Vizsgák kezdete (óra, pl. 8): ").strip()
            kezdes_ora = int(bev)
            if 0 <= kezdes_ora <= 23:
                break
            print("  Kérem 0-23 közötti értéket adjon meg.")
        except ValueError:
            print("  Kérem számot adjon meg.")

    while True:
        try:
            bev = input("  Vizsgák vége (óra, pl. 16): ").strip()
            veg_ora = int(bev)
            if veg_ora > kezdes_ora and veg_ora <= 24:
                break
            print(f"  Kérem {kezdes_ora}-nál nagyobb, max 24 értéket adjon meg.")
        except ValueError:
            print("  Kérem számot adjon meg.")

    napi_perc = (veg_ora - kezdes_ora) * 60
    max_vizsga = (napi_perc + SZUNET_HOSSZ) // (VIZSGA_HOSSZ + SZUNET_HOSSZ)

    print()
    print(f"  Napi időkeret:    {kezdes_ora}:00 – {veg_ora}:00 ({napi_perc} perc)")
    print(f"  Max vizsga/nap:   {max_vizsga}")
    print()

    return kezdes_ora, veg_ora, napi_perc


# ============================================================
# Belépési idő optimalizálás
# ============================================================

def optimalizal_belepesi_idok(esemenyek: list, vizsgalok: list):
    """
    Optimalizálja a vizsgázók belépési idejét.
    A vizsgázó az első vizsgája előtt 20 perccel (nyelvi tárgy esetén azonnal) lép be.
    """
    for v in vizsgalok:
        v_esem = [e for e in esemenyek if e.vizsgazo_nev == v.nev]
        if not v_esem:
            continue
        elso_esem = min(v_esem, key=lambda e: e.kezdes)
        prep = 0 if elso_esem.targy_nyelvi else 20
        v.belepesi_ido = elso_esem.kezdes - prep
        elso_esem.felkeszules_kezdet = v.belepesi_ido


# ============================================================
# Fő program
# ============================================================

def main():

    kezdes_ora, veg_ora, napi_perc = beolvas_idokeretet()

    # Vizsgázók beolvasása Excel fájlból
    while True:
        fajlnev = input("  Bemeneti Excel fájl neve (pl. vizsgalok.xlsx): ").strip()
        if not fajlnev:
            fajlnev = "vizsga_bemenet_sablon.xlsx"
            print(f"  Alapértelmezett: {fajlnev}")
        if not fajlnev.endswith(".xlsx"):
            fajlnev += ".xlsx"
        vizsgalok = beolvas_excelbol(fajlnev)
        if vizsgalok:
            break
        print("  Kérem adjon meg érvényes fájlt.")
        print()

    ossz_targy = sum(len(v.targyak) for v in vizsgalok)
    print()
    print(f"  Vizsgázók:    {len(vizsgalok)} fő")
    print(f"  Összes tárgy: {ossz_targy}")

    # Napokra osztás
    napok, max_vizsga_per_nap = napokra_oszt(vizsgalok, napi_perc)

    print(f"  Szükséges napok:  {len(napok)}")
    for i, nap in enumerate(napok, 1):
        nap_targyak = sum(len(v.targyak) for v in nap)
        print(f"    {i}. nap: {len(nap)} vizsgázó, {nap_targyak} vizsga"
              f" ({', '.join(v.nev for v in nap)})")
    print()

    # Napokra bontott szimuláció + kiírás
    print("=" * 80)
    print(f"              VIZSGA BEOSZTÁS — {len(napok)} NAPOS ÜTEMEZÉSI TERV")
    print(f"              Időkeret: {kezdes_ora}:00 – {veg_ora}:00"
          f" (max {max_vizsga_per_nap} vizsga/nap)")
    print("=" * 80)
    print()

    ossz_uresjarat = 0
    ossz_ido = 0
    ossz_max_var = 0
    nap_adatok = []  # Excel exporthoz

    for i, nap_vizsgalok in enumerate(napok, 1):
        print(f"  ── {i}. nap optimalizálása ──")
        print()
        esemenyek, nap_eredmeny = szimulacio(nap_vizsgalok)
        optimalizal_belepesi_idok(esemenyek, nap_eredmeny)
        ertek = _ertekeles(esemenyek, nap_eredmeny)
        ossz_ido += ertek[0]
        ossz_max_var = max(ossz_max_var, ertek[1])
        ossz_uresjarat += ertek[2]
        nap_adatok.append((esemenyek, nap_eredmeny, ertek))

        kiiras_nap(esemenyek, nap_eredmeny, i, kezdes_ora)

    # Összesítés
    if len(napok) > 1:
        print("=" * 80)
        print("  ÖSSZESÍTÉS")
        print("=" * 80)
        print(f"  Napok száma:       {len(napok)}")
        print(f"  Összes vizsgaidő:  {ossz_ido}p ({ossz_ido // 60}h {ossz_ido % 60}p)")
        print(f"  Max várakozás:     {ossz_max_var}p")
        print(f"  Összes üresjárat:  {ossz_uresjarat}p")
        print(f"  Összes kihasználtság: {ossz_targy * VIZSGA_HOSSZ / ossz_ido * 100:.1f}%")
        print()
        print("=" * 80)

    # Excel export
    fajlnev = excel_export(nap_adatok, kezdes_ora, veg_ora)
    if fajlnev:
        print()
        print(f"  ✔ Excel fájl elmentve: {fajlnev}")
        print()


if __name__ == "__main__":
    main()
